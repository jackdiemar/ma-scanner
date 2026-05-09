import pandas as pd
import numpy as np
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

def parse_amount(amount_str):
    """Convert amount string like '$100.00M' to float"""
    if pd.isna(amount_str) or amount_str == '':
        return np.nan
    
    amount_str = str(amount_str).strip()
    
    # Remove $ and commas
    amount_str = amount_str.replace('$', '').replace(',', '')
    
    # Handle millions (M) and billions (B)
    multiplier = 1
    if 'B' in amount_str.upper():
        multiplier = 1000
        amount_str = amount_str.upper().replace('B', '')
    elif 'M' in amount_str.upper():
        multiplier = 1
        amount_str = amount_str.upper().replace('M', '')
    elif 'K' in amount_str.upper():
        multiplier = 0.001
        amount_str = amount_str.upper().replace('K', '')
    
    try:
        return float(amount_str) * multiplier
    except:
        return np.nan

def is_vc_round(deal_type):
    """Pattern recognition to identify VC rounds"""
    if pd.isna(deal_type):
        return False
    
    deal_type_lower = str(deal_type).lower()
    
    vc_patterns = [
        r'seed',
        r'angel',
        r'early stage vc',
        r'later stage vc', 
        r'series [a-z]',
        r'series [a-z]\d',
        r'venture',
        r'convertible note',
        r'equity crowdfunding'
    ]
    
    for pattern in vc_patterns:
        if re.search(pattern, deal_type_lower):
            return True
    
    return False

def identify_exit_type(deal_type):
    """Identify if a deal is an exit and what type"""
    if pd.isna(deal_type):
        return None
    
    deal_type_lower = str(deal_type).lower()
    
    if re.search(r'ipo|initial public offering', deal_type_lower):
        return 'IPO'
    elif re.search(r'merger|acquisition|m&a', deal_type_lower):
        return 'M&A'
    elif re.search(r'buyout|lbo', deal_type_lower):
        return 'Buyout'
    else:
        return None

def format_excel_professionally(file_path, is_deal_history=True):
    """Apply professional formatting to Excel file"""
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Frozen header
    ws.freeze_panes = "A2"
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    if is_deal_history:
        col_widths = {
            'A': 15,  # PitchBook ID
            'B': 6,   # #
            'C': 35,  # Deal Type
            'D': 12,  # Date
            'E': 14,  # Amount
            'F': 16,  # Pre-Val
            'G': 16,  # Post-Val
            'H': 12,  # Status
            'I': 25,  # Stage
            'J': 10,  # Is_VC_Round
            'K': 18,  # Cumulative Total ($mm)
            'L': 18,  # Cumulative VC ($mm)
            'M': 50,  # Investor
            'N': 60   # Deal Synopsis
        }
        
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        # General info widths - auto-adjust based on content
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 16
    
    # Format currency columns (amounts are ALREADY in millions, just add formatting)
    currency_format = '$#,##0.0"M"'  # Shows as $X.XM (don't divide, already in millions)
    currency_cols = []
    
    for col_idx, cell in enumerate(ws[1], 1):
        col_letter = get_column_letter(col_idx)
        header_text = str(cell.value).lower() if cell.value else ""
        
        # Identify currency columns
        if any(keyword in header_text for keyword in ['amount', 'val', 'cumulative', 'raised', 'exit_amount']):
            currency_cols.append(col_letter)
    
    # Apply currency formatting to data rows
    for row in range(2, ws.max_row + 1):
        for col_letter in currency_cols:
            cell = ws[f'{col_letter}{row}']
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.0"M"'  # Already in millions
    
    # Format date columns
    for col_idx, cell in enumerate(ws[1], 1):
        col_letter = get_column_letter(col_idx)
        header_text = str(cell.value).lower() if cell.value else ""
        
        if 'date' in header_text:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value:
                    cell.number_format = 'YYYY-MM-DD'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # Data rows
                cell.alignment = Alignment(vertical='center')
    
    # Alternating row colors for readability
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = light_fill
    
    wb.save(file_path)
    print(f"   ✓ Applied professional formatting to {file_path}")

def analyze_deal_history(dh_file, general_file, output_dh_file, output_general_file):
    """Main analysis function with enhanced formatting"""
    
    print("=" * 80)
    print("DEAL HISTORY ANALYSIS & CLEANUP - ENHANCED VERSION")
    print("=" * 80)
    
    # Read files
    print("\n1. Reading files...")
    df_dh = pd.read_excel(dh_file, header=1)
    df_general = pd.read_excel(general_file, header=1)
    
    print(f"   ✓ Deal History: {len(df_dh)} rows")
    print(f"   ✓ General Info: {len(df_general)} companies")
    
    # Clean up columns
    print("\n2. Cleaning up columns...")
    
    meaningful_cols = ['PitchBook ID', '#', 'Deal Type', 'Date', 'Amount', 
                      'Pre-Val', 'Post-Val', 'Status', 'Stage', 'Investor', 'Deal Synopsis']
    
    cols_to_keep = [col for col in meaningful_cols if col in df_dh.columns]
    df_dh = df_dh[cols_to_keep].copy()
    
    print(f"   ✓ Reduced to {len(cols_to_keep)} meaningful columns")
    
    # Parse amounts
    print("\n3. Parsing monetary amounts...")
    df_dh['Amount_Parsed'] = df_dh['Amount'].apply(parse_amount)
    df_dh['PreVal_Parsed'] = df_dh['Pre-Val'].apply(parse_amount) if 'Pre-Val' in df_dh.columns else np.nan
    df_dh['PostVal_Parsed'] = df_dh['Post-Val'].apply(parse_amount) if 'Post-Val' in df_dh.columns else np.nan
    
    print(f"   ✓ Parsed {df_dh['Amount_Parsed'].notna().sum()} deal amounts")
    
    # Parse dates and sort
    print("\n4. Parsing and sorting by date...")
    df_dh['Date_Parsed'] = pd.to_datetime(df_dh['Date'], errors='coerce')
    df_dh = df_dh.sort_values(['PitchBook ID', 'Date_Parsed'], na_position='last')
    
    # Identify VC rounds
    print("\n5. Identifying VC rounds using pattern recognition...")
    df_dh['Is_VC_Round'] = df_dh['Deal Type'].apply(is_vc_round)
    
    vc_count = df_dh['Is_VC_Round'].sum()
    print(f"   ✓ Identified {vc_count} VC rounds out of {len(df_dh)} total deals")
    
    # Calculate cumulative funds raised
    print("\n6. Calculating cumulative funds raised...")
    
    def calculate_cumulative_by_group(group):
        """Calculate cumulative sum that forwards previous value when NaN"""
        group['Amount_For_Cumsum'] = group['Amount_Parsed'].fillna(0)
        group['Cumulative_Total_Raised'] = group['Amount_For_Cumsum'].cumsum()
        
        group['Amount_VC_For_Cumsum'] = group['Amount_Parsed'].where(group['Is_VC_Round'], 0).fillna(0)
        group['Cumulative_VC_Raised'] = group['Amount_VC_For_Cumsum'].cumsum()
        
        group = group.drop(['Amount_For_Cumsum', 'Amount_VC_For_Cumsum'], axis=1)
        
        return group
    
    df_dh = df_dh.groupby('PitchBook ID', group_keys=False).apply(calculate_cumulative_by_group)
    
    print(f"   ✓ Calculated cumulative totals for all companies")
    
    # Identify exits
    print("\n7. Identifying exits and updating general info...")
    
    df_dh['Exit_Type'] = df_dh['Deal Type'].apply(identify_exit_type)
    
    exit_info = []
    
    for pb_id in df_general['PitchBook ID']:
        company_deals = df_dh[df_dh['PitchBook ID'] == pb_id]
        exits = company_deals[company_deals['Exit_Type'].notna()]
        
        if len(exits) > 0:
            # Get the most recent exit
            latest_exit = exits.iloc[-1]
            
            # Get the status - IMPORTANT for cancelled/announced deals
            exit_status = latest_exit['Status'] if pd.notna(latest_exit['Status']) else 'Unknown'
            
            exit_info.append({
                'PitchBook ID': pb_id,
                'Has_Exit': 'Yes',
                'Exit_Type': latest_exit['Exit_Type'],
                'Exit_Status': exit_status,  # NEW: Completed, Cancelled, Announced, etc.
                'Exit_Date': latest_exit['Date'],
                'Exit_Amount_MM': latest_exit['Amount_Parsed'],
                'Exit_PostVal_MM': latest_exit['PostVal_Parsed'] if pd.notna(latest_exit.get('PostVal_Parsed')) else None
            })
        else:
            exit_info.append({
                'PitchBook ID': pb_id,
                'Has_Exit': 'No',
                'Exit_Type': None,
                'Exit_Status': None,
                'Exit_Date': None,
                'Exit_Amount_MM': None,
                'Exit_PostVal_MM': None
            })
    
    df_exit_info = pd.DataFrame(exit_info)
    df_general = df_general.merge(df_exit_info, on='PitchBook ID', how='left')
    
    exits_found = (df_general['Has_Exit'] == 'Yes').sum()
    completed_exits = (df_general['Exit_Status'] == 'Completed').sum()
    cancelled_exits = (df_general['Exit_Status'] == 'Cancelled').sum()
    announced_exits = (df_general['Exit_Status'] == 'Announced').sum()
    
    print(f"   ✓ Found exits for {exits_found} companies")
    print(f"      - Completed: {completed_exits}")
    print(f"      - Cancelled: {cancelled_exits}")
    print(f"      - Announced/Pending: {announced_exits}")
    
    # Prepare final Deal History with renamed columns
    print("\n8. Preparing final output with clear column names...")
    
    # Convert parsed amounts back to proper format for Excel
    df_dh['Amount_MM'] = df_dh['Amount_Parsed']
    df_dh['PreVal_MM'] = df_dh['PreVal_Parsed']
    df_dh['PostVal_MM'] = df_dh['PostVal_Parsed']
    
    final_dh_cols = {
        'PitchBook ID': 'PitchBook ID',
        '#': '#',
        'Deal Type': 'Deal Type',
        'Date': 'Date',
        'Amount_MM': 'Amount ($MM)',
        'PreVal_MM': 'Pre-Val ($MM)',
        'PostVal_MM': 'Post-Val ($MM)',
        'Status': 'Status',
        'Stage': 'Stage',
        'Is_VC_Round': 'Is VC Round?',
        'Cumulative_Total_Raised': 'Cumulative Total Raised ($MM)',
        'Cumulative_VC_Raised': 'Cumulative VC Raised ($MM)',
        'Investor': 'Investor',
        'Deal Synopsis': 'Deal Synopsis'
    }
    
    cols_available = [col for col in final_dh_cols.keys() if col in df_dh.columns]
    df_dh_final = df_dh[cols_available].copy()
    df_dh_final = df_dh_final.rename(columns=final_dh_cols)
    
    # Save to Excel
    print("\n9. Saving files with professional formatting...")
    
    df_dh_final.to_excel(output_dh_file, index=False, engine='openpyxl')
    df_general.to_excel(output_general_file, index=False, engine='openpyxl')
    
    # Apply formatting
    format_excel_professionally(output_dh_file, is_deal_history=True)
    format_excel_professionally(output_general_file, is_deal_history=False)
    
    print(f"   ✓ Saved: {output_dh_file}")
    print(f"   ✓ Saved: {output_general_file}")
    
    # Summary statistics
    print("\n" + "=" * 80)
    print("SUMMARY STATISTICS")
    print("=" * 80)
    
    print(f"\nDeal History:")
    print(f"  - Total deals: {len(df_dh_final):,}")
    print(f"  - VC rounds: {df_dh['Is_VC_Round'].sum():,}")
    print(f"  - Deals with amounts: {df_dh['Amount_Parsed'].notna().sum():,}")
    print(f"  - Total capital raised: ${df_dh['Amount_Parsed'].sum():,.1f}MM")
    print(f"  - Total VC capital: ${df_dh[df_dh['Is_VC_Round']]['Amount_Parsed'].sum():,.1f}MM")
    
    print(f"\nCompanies:")
    print(f"  - Total companies: {len(df_general):,}")
    print(f"  - Companies with exits: {exits_found}")
    print(f"  - Exit rate: {exits_found/len(df_general)*100:.1f}%")
    
    print("\n" + "=" * 80)
    print("✅ ANALYSIS COMPLETE WITH PROFESSIONAL FORMATTING!")
    print("=" * 80)

if __name__ == "__main__":
    dh_file = "combined_DH.xlsx"
    general_file = "general_info.xlsx"
    
    output_dh_file = "combined_DH_analyzed.xlsx"
    output_general_file = "general_info_updated.xlsx"
    
    analyze_deal_history(dh_file, general_file, output_dh_file, output_general_file)
